# 個人情報保護のため、一部をxxxxで表示してあります。
import openpyxl
import xlwings as xw
import os
import stat
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import subprocess
import shutil

# 前回のファイル名
former_filename = 'xxxx.xlsx'
# 本日のファイル名
today_filename = 'xxxx.xlsx'

# 共有フォルダから開くと解除できない読みとり専用になるので、いったん自分のフォルダに複製する。
def move_file():
    # shutil.copy(コピー元ファイルのパス,コピー先ファイルのパス)
    shutil.copy('./xxxx/' + former_filename, 'C:\\xxxx\\xxxx\\データ更新ファイル保留所\\' + former_filename)
    shutil.copy('./xxxx/' + today_filename, 'C:\\xxxx\\xxxx\\データ更新ファイル保留所\\' + today_filename)
move_file()

# 前回のファイルパス
before_file = '前回のファイルまでのパス' + former_filename
# 今回のファイルパス
thistime_file = '今回のファイルまでのパス' + today_filename


# 読み取り専用の解除
def editBook():
    try:
        # 読み取り専用を外す
        os.chmod(path=before_file, mode=stat.S_IWRITE)
        os.chmod(path=thistime_file, mode=stat.S_IWRITE)
        print('読み取り専用を解除しました。')
    except:
        print('読み取り専用解除が既に完了しています。')
editBook()

# ブックの参照
# xlwingでの処理
def move_sheet():
    f_before = xw.Book(before_file)
    t_thistime = xw.Book(thistime_file)
    # シートの移動(複製)
    f_before.sheets['自社版データ（参照用）'].copy(after=t_thistime.sheets[0])
    # フィルターの解除
    try:
        t_thistime.sheets['マスターファイル'].api.ShowAllData()
        print('フィルターを解除しました。')
    except:
        print('フィルターはかかっていません。')

    t_thistime.save(thistime_file)
    app = xw.apps.active
    app.quit()

move_sheet()

wb_before = openpyxl.load_workbook(before_file)
wb_thistime = openpyxl.load_workbook(thistime_file)

wb_master = wb_thistime['マスターファイル']
wb_refer = wb_thistime['自社版データ（参照用）']

# 漢字の〇を記号の○に置換
def repText():
    count = 0
    for row in wb_master.iter_rows():
        for cell in row:
            col_num = [16, 17, 18, 20]
            for num in range(0, len(col_num)):
                if cell.col_idx == col_num[num]:
                    if cell.value == '〇':
                        new_text = cell.value.replace("〇", "○")
                        cell.value = new_text
                        count += 1
                    elif cell.value == '0':
                        new_text = cell.value.replace("0", "○")
                        cell.value = new_text
                        count += 1
                    else:
                        pass
            

    print(str(count) + '件置換しました。')

    wb_thistime.save(thistime_file)

repText()

# vlookup処理 
def vLook():
    master_row = wb_master.max_row
    refer_row = wb_refer.max_row
    
    for i in range(2, master_row + 1):
        key_code = wb_master['H' + str(i)].value
        for k in range(3, refer_row + 1):
            eng_adress = wb_refer['H' + str(k)].value
            if key_code == eng_adress:
                mail_yn = wb_refer['Q' + str(k)].value
                if wb_master['R' + str(i)].value == '':
                    wb_master['R' + str(i)].value = mail_yn
                
                break
                
    wb_thistime.save(thistime_file)
    print('VLOOKUP処理が完了しました。')

vLook()

# js用セルの2行目までコピペし、最低価格2つの式を変更
def copyPaste():
    wb_before_master = wb_before['マスターファイル']
    col_name = ['AM', 'AN', 'AO', 'AP', 'AQ', 'AR']
    for i in range(1, 3):
        for cn in range(0, len(col_name)):
            wb_master[col_name[cn] + str(i)].value = wb_before_master[col_name[cn] + str(i)].value
        
    wb_master['AQ' + str(2)].value = wb_master['AP' + str(2)].value
    rowest = wb_master['AQ' + str(2)].value.replace('-', '/',1)
    wb_master['AQ' + str(2)].value = rowest
    wb_thistime.save(thistime_file)

copyPaste()

# 疑似オートフィル
def autoFill():
    master_row = wb_master.max_row
    col_name = ['AM', 'AN', 'AO', 'AP', 'AQ', 'AR']
    for n in range(3, master_row + 1):
        for cn in range(0, len(col_name)):
            # 1つ前の行の式が n行目のセルに入る。オートフィルではないので、そのまま1行前の数字が入る。
            wb_master[col_name[cn] + str(n)].value = wb_master[col_name[cn] + str(n-1)].value
            # 1つ前の行の数字をその行番号に変換する。
            change_v = wb_master[col_name[cn] + str(n)].value.replace(str(n-1), str(n))
            # 変換した数字に置換する。
            wb_master[col_name[cn] + str(n)].value = change_v
            # フォントを11でそろえる
            font1 = Font(size=11)
            wb_master[col_name[cn] + str(n)].font = font1
        # 価格を右揃え
        wb_master['AP' + str(n)].alignment = Alignment(horizontal="right")
        wb_master['AQ' + str(n)].alignment = Alignment(horizontal="right")
        wb_master['AR' + str(n)].alignment = Alignment(horizontal="right")
        
    wb_thistime.save(thistime_file)
    print('js用セルのオートフィルが完了しました。')

autoFill()
# 元の共有フォルダにファイルを戻す。
def reverseFile():
    rev_path = './EAJ更新ファイル/' + today_filename
    wb_thistime.save(rev_path)
reverseFile()

# 自分のフォルダにコピーしたファイルを削除
def removeFile():
    try:
        os.remove('C:\\xxxx\\xxxx\\データ更新保留所\\' + former_filename)
        os.remove('C:\\xxxx\\xxxx\\データ更新保留所\\' + today_filename)
        print('複製したファイルは削除しました。')
    except:
        print('複製ファイルの削除に失敗しました。')
removeFile()

# 更新したファイルを開く
def openThisTime():
    rev_path = './xxxx/' + today_filename
    subprocess.Popen(['start', rev_path], shell=True)

openThisTime()
